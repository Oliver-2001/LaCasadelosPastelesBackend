from flask import Flask, request, jsonify, send_file, render_template_string, make_response
from flask_jwt_extended import JWTManager, create_access_token
from flask_jwt_extended import jwt_required
from flask_jwt_extended import get_jwt_identity
from config import SQLALCHEMY_DATABASE_URI  # Importa la URI de conexión
from models import db, Usuario, Rol, Modulo, RolModulo, Producto, Inventario, Venta, DetalleVenta, Sucursal, PrediccionesIA # Importa tu base de datos y modelo de usuario
from werkzeug.security import check_password_hash, generate_password_hash
from flask_cors import CORS
from datetime import datetime, timedelta, date
from sqlalchemy import func, text
from openpyxl import Workbook
from weasyprint import HTML
import io
from reportlab.lib.pagesizes import letter
from io import BytesIO
from reportlab.pdfgen import canvas
from reportlab.lib.units import inch
from reportlab.lib import colors
from reportlab.platypus import Table, TableStyle, SimpleDocTemplate, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet
from predicciones import generar_predicciones

app = Flask(__name__)
CORS(app, origins="http://localhost:3000")

# Configuración de la base de datos
app.config["SQLALCHEMY_DATABASE_URI"] = SQLALCHEMY_DATABASE_URI
app.config["SQLALCHEMY_TRACK_MODIFICATIONS"] = False
app.config["SECRET_KEY"] = "Or005836434*"
app.config["JWT_SECRET_KEY"] = "OR005836434*"

# Inicialización de la base de datos y JWT
db.init_app(app)
jwt = JWTManager(app)

###################################################################################Rutas#############################################################

@app.route("/")
def home():
    return jsonify(message="Bienvenido a la API de la pastelería"), 200

# ############################################# Ruta de inicio de sesión #############################################
@app.route('/login', methods=['POST'])
def login():
    data = request.get_json()
    usuario = data.get('usuario')
    contrasena = data.get('contrasena')

    user = Usuario.query.filter_by(usuario=usuario).first()

    if user and check_password_hash(user.contrasena, contrasena):
        # ✅ Genera el token JWT
        token = create_access_token(identity=str(user.id_usuario))  # Guarda el ID del usuario en el token
        return jsonify({
            "message": "Login exitoso",
            "token": token
        }), 200
    else:
        return jsonify({"message": "Usuario o contraseña incorrectos"}), 401
    

############################################## Ruta protegida (requiere autenticación)#############################################
@app.route("/protected", methods=["GET"])
@jwt_required()
def protected():
    current_user = get_jwt_identity()
    return jsonify(message=f"Hola, {current_user}! Esta es una ruta protegida."), 200

# Ruta para obtener los modulos del usuario autenticado
@app.route('/modulos', methods=['GET'])
@jwt_required()
def obtener_modulos():
    current_user_id = get_jwt_identity() 
    usuario = Usuario.query.filter_by(id_usuario=current_user_id).first()  

    if not usuario:
        return jsonify({"message": "Usuario no encontrado"}), 404

    # Obtener los módulos a los que el usuario tiene acceso
    modulos = Modulo.query.join(RolModulo).filter(RolModulo.id_rol == usuario.id_rol).all()

    modulos_list = [{"id_modulo": modulo.id_modulo, "nombre": modulo.nombre, "descripcion": modulo.descripcion} for modulo in modulos]

    return jsonify(modulos_list), 200


############################################## Ruta para obtener todos los usuarios (solo accesible por admins) (Modulo Admin de Usuarios) #############################################
@app.route('/usuarios', methods=['GET'])
@jwt_required()
def obtener_usuarios():
    current_user_id = get_jwt_identity()
    usuario = Usuario.query.filter_by(id_usuario=current_user_id).first()

    if not usuario:
        return jsonify({"message": "Usuario no encontrado"}), 404

    if usuario.id_rol != 1:  
        return jsonify({"message": "No tienes permisos para acceder a esta información"}), 403

    # Obtener todos los usuarios 
    usuarios = Usuario.query.all()
    usuarios_list = [{"id_usuario": u.id_usuario, "nombre": u.nombre, "usuario": u.usuario, "rol": u.id_rol} for u in usuarios]

    return jsonify(usuarios_list), 200

################################################## Ruta para eliminar usuarios (Modulo Admin de Usuarios) #############################################
@app.route('/usuarios/<int:id_usuario>', methods=['DELETE'])
@jwt_required()
def eliminar_usuario(id_usuario):
    current_user_id = get_jwt_identity()
    usuario_actual = Usuario.query.filter_by(id_usuario=current_user_id).first()

    if not usuario_actual:
        return jsonify({"message": "Usuario no encontrado"}), 404

    if usuario_actual.id_rol != 1:  
        return jsonify({"message": "No tienes permisos para eliminar usuarios"}), 403

    usuario = Usuario.query.filter_by(id_usuario=id_usuario).first()

    if not usuario:
        return jsonify({"message": "El usuario no existe"}), 404

    if usuario.id_rol == 1:
        return jsonify({"message": "No se puede eliminar al Superadmin"}), 403

    db.session.delete(usuario)
    db.session.commit()

    return jsonify({"message": "Usuario eliminado correctamente"}), 200


################################# Ruta para editar usuarios (Modulo Admin de Usuarios) #############################################
@app.route('/usuarios/<int:id_usuario>', methods=['PUT'])
@jwt_required()
def editar_usuario(id_usuario):
    current_user_id = get_jwt_identity()
    usuario_actual = Usuario.query.filter_by(id_usuario=current_user_id).first()

    if not usuario_actual:
        return jsonify({"message": "Usuario no encontrado"}), 404

    if usuario_actual.id_rol != 1:
        return jsonify({"message": "No tienes permisos para modificar usuarios"}), 403

    usuario = Usuario.query.filter_by(id_usuario=id_usuario).first()
    if not usuario:
        return jsonify({"message": "Usuario no encontrado"}), 404

    data = request.get_json()
    nombre = data.get("nombre")
    usuario_nombre = data.get("usuario")
    id_rol = data.get("id_rol")

    if not nombre or not usuario_nombre or not id_rol:
        return jsonify({"message": "Todos los campos son obligatorios"}), 400

    if id_rol not in [1, 2, 3, 4]:
        return jsonify({"message": "Rol no válido"}), 400

    usuario_existente = Usuario.query.filter_by(usuario=usuario_nombre).first()
    if usuario_existente and usuario_existente.id_usuario != id_usuario:
        return jsonify({"message": "El nombre de usuario ya está en uso"}), 400

    usuario.nombre = nombre
    usuario.usuario = usuario_nombre
    usuario.id_rol = id_rol

    db.session.commit()

    return jsonify({"message": "Usuario actualizado correctamente"}), 200

######################################### Ruta para crear usuarios nuevos (Modulo Admin de Usuarios) #############################################
@app.route('/usuarios', methods=['POST'])
def crear_usuario():
    data = request.get_json()

    print("Datos recibidos:", data)  # Depuración: Ver lo que se está recibiendo

    # Validar que se envíen todos los campos
    if not data.get('nombre') or not data.get('usuario') or not data.get('contrasena') or not data.get('id_rol'):
        return jsonify({'error': 'Faltan datos obligatorios'}), 400

    # Verificar que el id_rol sea válido
    rol = Rol.query.get(data['id_rol'])
    if not rol:
        return jsonify({'error': 'Rol no válido'}), 400

    # Crear el nuevo usuario
    nuevo_usuario = Usuario(
        nombre=data['nombre'],
        usuario=data['usuario'],
        contrasena=generate_password_hash(data['contrasena']),
        id_rol=data['id_rol']
    )

    try:
        db.session.add(nuevo_usuario)
        db.session.commit()
        return jsonify({
            'mensaje': 'Usuario creado con éxito',
            'usuario': {
                'id_usuario': nuevo_usuario.id_usuario,
                'nombre': nuevo_usuario.nombre,
                'usuario': nuevo_usuario.usuario,
                'rol': rol.nombre
            }
        }), 201
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 500


################################################################## Modulo Productos ##############################################################
@app.route('/productos', methods=['GET'])
@jwt_required()
def obtener_productos():
    current_user_id = get_jwt_identity()

    usuario = Usuario.query.filter_by(id_usuario=current_user_id).first()

    if not usuario:
        return jsonify({"message": "Usuario no encontrado"}), 404

    # Obtener todos los productos
    productos = Producto.query.all() 

    productos_list = [
        {
            "id_producto": p.id_producto,
            "nombre": p.nombre,
            "precio": p.precio,
            "stock": p.stock,
            "categoria": p.categoria
        } for p in productos
    ]

    return jsonify(productos_list), 200



##############(EDITAR) Modulo Productos #############################


@app.route('/productos/<int:id_producto>', methods=['PUT'])
@jwt_required()
def editar_producto(id_producto):
    current_user_id = get_jwt_identity()
    usuario = Usuario.query.filter_by(id_usuario=current_user_id).first()

    if not usuario:
        return jsonify({"message": "Usuario no encontrado"}), 404

    if usuario.rol.nombre not in ['superadmin', 'admin']:  
        return jsonify({"message": "No tiene permisos para editar productos"}), 403

    producto = Producto.query.get(id_producto)
    if not producto:
        return jsonify({"message": "Producto no encontrado"}), 404

    # Obtener los datos enviados en el cuerpo de la solicitud
    data = request.get_json()

    # Actualizar los campos del producto
    producto.nombre = data.get('nombre', producto.nombre)
    producto.precio = data.get('precio', producto.precio)
    producto.stock = data.get('stock', producto.stock)
    producto.categoria = data.get('categoria', producto.categoria)

    db.session.commit()

    return jsonify({"message": "Producto actualizado correctamente"}), 200


##############(Eliminar) Modulo Productos ########################################

@app.route('/productos/<int:id_producto>', methods=['DELETE'])
@jwt_required()
def eliminar_producto(id_producto):
    current_user_id = get_jwt_identity()
    usuario = Usuario.query.filter_by(id_usuario=current_user_id).first()

    if not usuario:
        return jsonify({"message": "Usuario no encontrado"}), 404

    # Accede a 'nombre_rol' ya que tu clase Rol tiene esa columna
    if usuario.rol.nombre not in ['superadmin', 'admin']:  
        return jsonify({"message": "No tiene permisos para eliminar productos"}), 403

    producto = Producto.query.get(id_producto)
    if not producto:
        return jsonify({"message": "Producto no encontrado"}), 404

    db.session.delete(producto)
    db.session.commit()

    return jsonify({"message": "Producto eliminado correctamente"}), 200

############################## (Crear) Modulo Productos ########################################

@app.route('/productos', methods=['POST'])
@jwt_required()
def crear_producto():
    current_user_id = get_jwt_identity()
    usuario = Usuario.query.filter_by(id_usuario=current_user_id).first()

    if not usuario:
        return jsonify({"message": "Usuario no encontrado"}), 404

    if usuario.rol.nombre not in ['superadmin', 'admin']:
        return jsonify({"message": "No tiene permisos para crear productos"}), 403

    data = request.get_json()

    nombre = data.get('nombre')
    precio = data.get('precio')
    stock = data.get('stock')
    categoria = data.get('categoria')

    if not nombre or precio is None or stock is None or not categoria:
        return jsonify({"message": "Faltan datos obligatorios"}), 400

    # Verificar si ya existe un producto con el mismo nombre
    producto_existente = Producto.query.filter_by(nombre=nombre).first()
    if producto_existente:
        return jsonify({"message": "Ya existe un producto con ese nombre"}), 409

    nuevo_producto = Producto(
        nombre=nombre,
        precio=precio,
        stock=stock,
        categoria=categoria
    )

    try:
        db.session.add(nuevo_producto)
        db.session.commit()

        return jsonify({
            "message": "Producto creado correctamente",
            "producto": {
                "id_producto": nuevo_producto.id_producto,
                "nombre": nuevo_producto.nombre,
                "precio": nuevo_producto.precio,
                "stock": nuevo_producto.stock,
                "categoria": nuevo_producto.categoria
            }
        }), 201

    except Exception as e:
        db.session.rollback()
        return jsonify({"error": str(e)}), 500



################################################################## Modulo Inventario ##############################################################


@app.route('/inventario', methods=['GET'])
@jwt_required()
def obtener_inventario():
    # Obtén el ID del usuario autenticado
    current_user_id = get_jwt_identity()
    usuario = Usuario.query.filter_by(id_usuario=current_user_id).first()

    if not usuario:
        return jsonify({"message": "Usuario no encontrado"}), 404

    inventarios = Inventario.query.all()
    inventarios_list = [
        {
            "id_insumo": inventario.id_insumo,  # Agregar el id_insumo
            "nombre": inventario.nombre,
            "cantidad": inventario.cantidad,
            "unidad": inventario.unidad,
            "fecha_actualizacion": inventario.fecha_actualizacion,
            "id_sucursal": inventario.id_sucursal
        } for inventario in inventarios
    ]

    return jsonify(inventarios_list), 200

################## (Editar) Modulo Inventario ########################################

@app.route('/inventario/<int:id_insumo>', methods=['PUT'])
@jwt_required()
def editar_inventario(id_insumo):
    data = request.get_json()

    inventario = Inventario.query.get(id_insumo)
    if not inventario:
        return jsonify({"message": "Insumo no encontrado"}), 404

    # Validar y actualizar campos
    inventario.nombre = data.get("nombre", inventario.nombre)
    inventario.cantidad = data.get("cantidad", inventario.cantidad)
    inventario.unidad = data.get("unidad", inventario.unidad)
    inventario.fecha_actualizacion = datetime.utcnow()  # Actualizar la fecha

    db.session.commit()
    return jsonify({"message": "Insumo actualizado correctamente."}), 200


################## (Eliminar) Modulo Inventario ########################################

@app.route('/inventario/<int:id_insumo>', methods=['DELETE'])
@jwt_required()
def eliminar_inventario(id_insumo):
    inventario = Inventario.query.get(id_insumo)
    if not inventario:
        return jsonify({"message": "Insumo no encontrado"}), 404

    db.session.delete(inventario)
    db.session.commit()

    return jsonify({"message": "Insumo eliminado correctamente."}), 200

#################### (Crear) Modulo Inventario ########################################

@app.route('/inventario', methods=['POST'])
@jwt_required()
def crear_insumo():
    current_user_id = get_jwt_identity()
    usuario = Usuario.query.filter_by(id_usuario=current_user_id).first()

    if not usuario:
        return jsonify({"message": "Usuario no encontrado"}), 404

    data = request.get_json()
    nombre = data.get("nombre")
    cantidad = data.get("cantidad")
    unidad = data.get("unidad")
    id_sucursal = data.get("id_sucursal")

    # Validaciones básicas
    if not all([nombre, cantidad, unidad, id_sucursal]):
        return jsonify({"message": "Faltan datos requeridos"}), 400

    # Verificar si ya existe un insumo con el mismo nombre en esa sucursal
    insumo_existente = Inventario.query.filter_by(nombre=nombre, id_sucursal=id_sucursal).first()
    if insumo_existente:
        return jsonify({"message": f"El insumo '{nombre}' ya existe en esta sucursal"}), 409

    nuevo_insumo = Inventario(
        nombre=nombre,
        cantidad=cantidad,
        unidad=unidad,
        id_sucursal=id_sucursal,
        fecha_actualizacion=datetime.utcnow()
    )

    db.session.add(nuevo_insumo)
    db.session.commit()

    return jsonify({"message": "Insumo creado exitosamente"}), 201

############################################ (Crear) Modulo VENTAS #####################################################

@app.route('/ventas', methods=['POST'])
@jwt_required()
def crear_venta():
    try:
        data = request.json
        detalles = data.get('detalles', [])
        id_usuario = get_jwt_identity()

        if not detalles or not id_usuario:
            return jsonify({'mensaje': 'Faltan datos obligatorios (detalles o id_usuario).'}), 400

        total = 0
        detalles_venta = []

        for item in detalles:
            id_producto = item['id_producto']
            cantidad = item['cantidad']

            producto = Producto.query.get(id_producto)
            if not producto:
                return jsonify({'mensaje': f'Producto con id {id_producto} no encontrado.'}), 404

            if producto.stock < cantidad:
                return jsonify({'mensaje': f'Stock insuficiente para producto {producto.nombre}.'}), 400

            subtotal = producto.precio * cantidad
            total += subtotal

            # Preparar detalle
            detalle = DetalleVenta(
                id_producto=id_producto,
                cantidad=cantidad,
                subtotal=subtotal
            )
            detalles_venta.append(detalle)

            # Actualizar stock
            producto.stock -= cantidad

        # Crear la venta
        venta = Venta(
            fecha=datetime.now(),
            total=total,
            id_usuario=id_usuario
        )
        db.session.add(venta)
        db.session.flush()  # Para obtener el id_venta antes de hacer commit

        # Asociar detalles a la venta
        for detalle in detalles_venta:
            detalle.id_venta = venta.id_venta
            db.session.add(detalle)

        db.session.commit()
        return jsonify({'mensaje': 'Venta registrada correctamente', 'id_venta': venta.id_venta}), 201

    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 500
    
 ############################################ (Crear PDF DE LA VENTA) Modulo VENTAS #####################################################

@app.route('/ventas/<int:id_venta>/reporte-pdf', methods=['GET'])
def reporte_pdf_venta(id_venta):
    venta = Venta.query.get(id_venta)
    if not venta:
        return jsonify({'error': 'Venta no encontrada'}), 404

    usuario = Usuario.query.get(venta.id_usuario)
    detalles = (
        db.session.query(
            Producto.nombre,
            DetalleVenta.cantidad,
            Producto.precio,
            DetalleVenta.subtotal
        )
        .join(Producto, Producto.id_producto == DetalleVenta.id_producto)
        .filter(DetalleVenta.id_venta == id_venta)
        .all()
    )

    html_template = """
    <html>
    <head>
        <style>
            body { font-family: sans-serif; }
            table { border-collapse: collapse; width: 100%; margin-top: 20px;}
            th, td { border: 1px solid #333; padding: 8px; text-align: left; }
            th { background-color: #f2f2f2; }
            h2, h3 { margin-bottom: 5px; }
        </style>
    </head>
    <body>
        <h2>Reporte de Venta</h2>
        <p><strong>Fecha y hora:</strong> {{ venta.fecha.strftime('%Y-%m-%d %H:%M:%S') }}</p>
        <p><strong>Registrado por:</strong> {{ usuario.nombre if usuario else 'Desconocido' }}</p>

        <table>
            <thead>
                <tr>
                    <th>Producto</th>
                    <th>Cantidad</th>
                    <th>Precio Unitario (Q)</th>
                    <th>Subtotal (Q)</th>
                </tr>
            </thead>
            <tbody>
                {% for nombre, cantidad, precio, subtotal in detalles %}
                <tr>
                    <td>{{ nombre }}</td>
                    <td>{{ cantidad }}</td>
                    <td>{{ "%.2f"|format(precio) }}</td>
                    <td>{{ "%.2f"|format(subtotal) }}</td>
                </tr>
                {% endfor %}
                <tr>
                    <td colspan="3" style="text-align: right; font-weight: bold;">Total</td>
                    <td>{{ "%.2f"|format(venta.total) }}</td>
                </tr>
            </tbody>
        </table>
    </body>
    </html>
    """

    html_rendered = render_template_string(
        html_template,
        venta=venta,
        usuario=usuario,
        detalles=detalles
    )

    pdf = HTML(string=html_rendered).write_pdf()

    response = make_response(pdf)
    response.headers['Content-Type'] = 'application/pdf'
    response.headers['Content-Disposition'] = f'attachment; filename=reporte_venta_{id_venta}.pdf'
    return response



############################################ LISTAR VENTAS #####################################################
@app.route('/ventas', methods=['GET'])
def listar_ventas():
    try:
        # Parámetros opcionales
        fecha_str = request.args.get('fecha')
        id_usuario = request.args.get('id_usuario')

        query = Venta.query

        if fecha_str:
            try:
                fecha = datetime.strptime(fecha_str, '%Y-%m-%d').date()
                query = query.filter(db.func.cast(Venta.fecha, db.Date) == fecha)
            except ValueError:
                return jsonify({'error': 'Formato de fecha inválido. Usa YYYY-MM-DD'}), 400

        if id_usuario:
            query = query.filter(Venta.id_usuario == int(id_usuario))

        ventas = query.all()

        resultado = [
            {
                'id_venta': venta.id_venta,
                'fecha': venta.fecha.strftime('%Y-%m-%d %H:%M:%S'),
                'total': venta.total,
                'id_usuario': venta.id_usuario
            }
            for venta in ventas
        ]

        return jsonify(resultado), 200

    except Exception as e:
        return jsonify({'error': str(e)}), 500
    
############################################ LISTAR DETALLES DE VENTA #####################################################
@app.route('/ventas/<int:id_venta>', methods=['GET'])
def obtener_detalle_venta(id_venta):
    try:
        venta = Venta.query.get(id_venta)
        if not venta:
            return jsonify({'mensaje': f'No se encontró la venta con ID {id_venta}'}), 404

        detalles = DetalleVenta.query.filter_by(id_venta=id_venta).all()

        detalle_items = []
        for detalle in detalles:
            producto = Producto.query.get(detalle.id_producto)
            detalle_items.append({
                'id_producto': detalle.id_producto,
                'nombre_producto': producto.nombre if producto else 'Desconocido',
                'cantidad': detalle.cantidad,
                'subtotal': detalle.subtotal
            })

        resultado = {
            'id_venta': venta.id_venta,
            'fecha': venta.fecha.strftime('%Y-%m-%d %H:%M:%S'),
            'total': venta.total,
            'id_usuario': venta.id_usuario,
            'detalles': detalle_items
        }

        return jsonify(resultado), 200

    except Exception as e:
        return jsonify({'error': str(e)}), 500
    
############################################## Reporte de ventas del día (VENTAS)#####################################################

@app.route("/reporte-ventas-dia", methods=["GET"])
@jwt_required()
def reporte_ventas_dia():
    hoy = date.today()
    inicio_dia = datetime.combine(hoy, datetime.min.time())
    fin_dia = datetime.combine(hoy, datetime.max.time())

    ventas = Venta.query.filter(Venta.fecha.between(inicio_dia, fin_dia)).all()
    if not ventas:
        return jsonify({"message": "No hay ventas registradas hoy."}), 404

    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter)
    elements = []
    styles = getSampleStyleSheet()

    # Título
    title = Paragraph("📋 <b>Reporte de Ventas del Día</b>", styles['Title'])
    date_info = Paragraph(f"📅 Fecha: {hoy.strftime('%d/%m/%Y')}", styles['Normal'])
    elements.extend([title, Spacer(1, 12), date_info, Spacer(1, 20)])

    # Encabezado de la tabla
    data = [["Fecha", "Producto", "Cantidad", "Subtotal (Q)"]]
    total_general = 0

    for venta in ventas:
        detalles = DetalleVenta.query.filter_by(id_venta=venta.id_venta).all()
        for detalle in detalles:
            producto = Producto.query.get(detalle.id_producto)
            fila = [
                venta.fecha.strftime("%d/%m/%Y %H:%M"),
                producto.nombre,
                detalle.cantidad,
                f"Q{detalle.subtotal:.2f}"
            ]
            data.append(fila)
            total_general += detalle.subtotal

    # Estilo de tabla
    tabla = Table(data, colWidths=[120, 200, 80, 100])
    tabla.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.orange),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 10),
        ("BOTTOMPADDING", (0, 0), (-1, 0), 8),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.black),
    ]))

    elements.append(tabla)
    elements.append(Spacer(1, 20))

    # Total
    total_text = Paragraph(f"<b>TOTAL VENDIDO HOY:</b> Q{total_general:.2f}", styles["Heading2"])
    elements.append(total_text)

    doc.build(elements)
    buffer.seek(0)

    return send_file(
        buffer,
        as_attachment=True,
        download_name="reporte_ventas_dia.pdf",
        mimetype="application/pdf"
    )   


################################################################## Modulo Sucursales ##############################################################

@app.route('/sucursales', methods=['GET'])
def obtener_sucursales():
    sucursales = Sucursal.query.all()
    resultado = []
    for sucursal in sucursales:
        resultado.append({
            'id_sucursal': sucursal.id_sucursal,
            'nombre': sucursal.nombre,
            'direccion': sucursal.direccion,
            'latitud': sucursal.latitud,
            'longitud': sucursal.longitud
        })
    return jsonify(resultado)

############################################ (Crear) Modulo Sucursales #####################################################
@app.route('/sucursales', methods=['POST'])
def agregar_sucursal():
    data = request.get_json()

    nombre = data.get('nombre')
    direccion = data.get('direccion')
    latitud = data.get('latitud')
    longitud = data.get('longitud')

    if not all([nombre, direccion, latitud, longitud]):
        return jsonify({'error': 'Faltan datos requeridos'}), 400

    try:
        latitud = float(latitud)
        longitud = float(longitud)
    except ValueError:
        return jsonify({'error': 'Latitud y longitud deben ser numéricos'}), 400

    nueva_sucursal = Sucursal(
        nombre=nombre,
        direccion=direccion,
        latitud=latitud,
        longitud=longitud
    )

    db.session.add(nueva_sucursal)
    db.session.commit()

    return jsonify({'message': 'Sucursal agregada exitosamente'}), 201

############################################ (Eliminar) Modulo Sucursales #####################################################

@app.route('/sucursales/<int:id_sucursal>', methods=['DELETE'])
def eliminar_sucursal(id_sucursal):
    sucursal = Sucursal.query.get(id_sucursal)

    if not sucursal:
        return jsonify({'error': 'Sucursal no encontrada'}), 404

    try:
        db.session.delete(sucursal)
        db.session.commit()
        return jsonify({'message': 'Sucursal eliminada exitosamente'}), 200
    except Exception as e:
        db.session.rollback()
        print('Error al eliminar sucursal:', e)
        return jsonify({'error': 'No se pudo eliminar la sucursal'}), 500


############################################ Modulo IA #####################################################

@app.route("/generar_predicciones", methods=["POST"])
def generar_predicciones_endpoint():
    resultado = generar_predicciones()
    return jsonify({"mensaje": resultado}), 200

############################################## Modulo GET PREDICCIONES #####################################################

@app.route('/predicciones', methods=['GET'])
def obtener_predicciones():
    """
    Endpoint para obtener predicciones filtradas por:
    - id_producto (opcional)
    - id_sucursal (opcional, default 1)
    - fecha_inicio (opcional, formato YYYY-MM-DD)
    - fecha_fin (opcional, formato YYYY-MM-DD)
    """
    id_producto = request.args.get('id_producto', type=int)
    id_sucursal = request.args.get('id_sucursal', default=1, type=int)
    fecha_inicio = request.args.get('fecha_inicio')
    fecha_fin = request.args.get('fecha_fin')

    query = (
    db.session.query(
        PrediccionesIA.id_producto,
        Producto.nombre.label("nombre_producto"),
        PrediccionesIA.fecha_prediccion,
        PrediccionesIA.cantidad_prediccion,
        PrediccionesIA.id_sucursal
    )
    .join(Producto, Producto.id_producto == PrediccionesIA.id_producto)
    .filter(PrediccionesIA.id_sucursal == id_sucursal)
)
    if id_producto:
        query = query.filter(PrediccionesIA.id_producto == id_producto)

    if fecha_inicio:
        try:
            fecha_inicio_dt = datetime.strptime(fecha_inicio, '%Y-%m-%d')
            query = query.filter(PrediccionesIA.fecha_prediccion >= fecha_inicio_dt)
        except ValueError:
            return jsonify({"error": "Formato de fecha_inicio inválido, debe ser YYYY-MM-DD"}), 400

    if fecha_fin:
        try:
            fecha_fin_dt = datetime.strptime(fecha_fin, '%Y-%m-%d')
            query = query.filter(PrediccionesIA.fecha_prediccion <= fecha_fin_dt)
        except ValueError:
            return jsonify({"error": "Formato de fecha_fin inválido, debe ser YYYY-MM-DD"}), 400

    predicciones = query.order_by(PrediccionesIA.fecha_prediccion).all()

    resultado = []
    for p in predicciones:
        resultado.append({
        "id_producto": p.id_producto,
        "nombre_producto": p.nombre_producto,
        "fecha_prediccion": p.fecha_prediccion.strftime('%Y-%m-%d'),
        "cantidad_prediccion": float(p.cantidad_prediccion) if p.cantidad_prediccion is not None else 0,
        "id_sucursal": p.id_sucursal
    })



    return jsonify(resultado)


############################################## Modulo INICIO #####################################################

@app.route('/ventas/por-producto', methods=['GET'])
def ventas_por_producto():
    fecha_limite = datetime.now() - timedelta(days=30)
    # Join entre DetalleVenta, Venta y Producto
    resultados = (
        db.session.query(
            Producto.id_producto,
            Producto.nombre,
            func.sum(DetalleVenta.cantidad).label('cantidad_vendida')
        )
        .join(DetalleVenta, Producto.id_producto == DetalleVenta.id_producto)
        .join(Venta, DetalleVenta.id_venta == Venta.id_venta)
        .filter(Venta.fecha >= fecha_limite)
        .group_by(Producto.id_producto, Producto.nombre)
        .order_by(func.sum(DetalleVenta.cantidad).desc())
        .all()
    )

    respuesta = []
    for id_producto, nombre, cantidad_vendida in resultados:
        respuesta.append({
            'id_producto': id_producto,
            'nombre_producto': nombre,
            'cantidad': cantidad_vendida
        })

    return jsonify(respuesta)

############################################### Modulo INICIO #####################################################

@app.route('/ventas/diarias', methods=['GET'])
def ventas_diarias():
    fecha_limite = datetime.now() - timedelta(days=30)

    resultados = (
        db.session.query(
            Venta.fecha,
            func.sum(Venta.total).label('total_ventas')
        )
        .filter(Venta.fecha >= fecha_limite)
        .group_by(Venta.fecha)
        .order_by(Venta.fecha.asc())
        .all()
    )

    respuesta = []
    for fecha, total_ventas in resultados:
        respuesta.append({
            'fecha': fecha.strftime('%Y-%m-%d'),
            'total_ventas': float(total_ventas)
        })

    return jsonify(respuesta)

############################################### Modulo INICIO #####################################################

@app.route('/productos/mas-vendidos', methods=['GET'])
def productos_mas_vendidos():
    sql = text("""
        SELECT p.nombre, SUM(dv.cantidad) AS cantidad_vendida
        FROM DetallesVenta dv
        JOIN Productos p ON dv.id_producto = p.id_producto
        GROUP BY p.nombre
        ORDER BY cantidad_vendida DESC
    """)
    result = db.session.execute(sql).fetchall()

    productos = []
    for row in result:
        productos.append({
            "nombre": row.nombre,
            "cantidad_vendida": int(row.cantidad_vendida)
        })

    return {"productos": productos}


################################################## MODULO REPORTES #####################################################

@app.route('/reporte/excel/ventas-producto', methods=['GET'])
def reporte_excel_ventas_producto():
    fecha_limite = datetime.now() - timedelta(days=30)

    resultados = (
        db.session.query(
            Producto.nombre,
            func.sum(DetalleVenta.cantidad).label('cantidad_vendida')
        )
        .join(DetalleVenta, Producto.id_producto == DetalleVenta.id_producto)
        .join(Venta, DetalleVenta.id_venta == Venta.id_venta)
        .filter(Venta.fecha >= fecha_limite)
        .group_by(Producto.nombre)
        .order_by(func.sum(DetalleVenta.cantidad).desc())
        .all()
    )

    wb = Workbook()
    ws = wb.active
    ws.title = "Ventas por Producto"

    # Encabezados
    ws.append(["Producto", "Cantidad Vendida"])

    # Datos
    for nombre, cantidad in resultados:
        ws.append([nombre, cantidad])

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(
        output,
        as_attachment=True,
        download_name="reporte_ventas_productos.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

############################################### Modulo Reportes #####################################################
@app.route('/reporte/pdf/ventas-detalladas-diarias', methods=['GET'])
def reporte_pdf_ventas_detalladas_diarias():
    fecha_limite = datetime.now() - timedelta(days=30)

    # Primero, obtenemos las fechas de ventas del último mes
    fechas = (
        db.session.query(Venta.fecha)
        .filter(Venta.fecha >= fecha_limite)
        .group_by(Venta.fecha)
        .order_by(Venta.fecha.asc())
        .all()
    )
    fechas = [f[0] for f in fechas]  # lista de fechas

    datos_por_fecha = []

    for fecha in fechas:
        # Por cada fecha, obtener productos vendidos, cantidad y subtotal
        detalles = (
            db.session.query(
                Producto.nombre,
                func.sum(DetalleVenta.cantidad).label('cantidad_vendida'),
                func.sum(DetalleVenta.cantidad * Producto.precio).label('total_producto')
            )
            .join(DetalleVenta, Producto.id_producto == DetalleVenta.id_producto)
            .join(Venta, DetalleVenta.id_venta == Venta.id_venta)
            .filter(Venta.fecha == fecha)
            .group_by(Producto.nombre)
            .all()
        )

        total_dia = sum(d.total_producto for d in detalles if d.total_producto is not None)

        datos_por_fecha.append({
            'fecha': fecha,
            'detalles': detalles,
            'total_dia': total_dia
        })

    # Plantilla HTML con Jinja2
    html_template = """
    <html>
    <head>
        <style>
            body { font-family: sans-serif; }
            table { border-collapse: collapse; width: 100%; margin-bottom: 30px;}
            th, td { border: 1px solid #333; padding: 5px; text-align: left; }
            th { background-color: #f2f2f2; }
            h2 { margin-bottom: 5px; }
            .total { font-weight: bold; }
        </style>
    </head>
    <body>
        <h1>Reporte Detallado de Ventas Diarias (Últimos 30 días)</h1>
        {% for dia in datos %}
            <h2>Fecha: {{ dia.fecha.strftime('%Y-%m-%d') }}</h2>
            <table>
                <thead>
                    <tr>
                        <th>Producto</th>
                        <th>Cantidad Vendida</th>
                        <th>Total (Q)</th>
                    </tr>
                </thead>
                <tbody>
                    {% for producto, cantidad, total in dia.detalles %}
                    <tr>
                        <td>{{ producto }}</td>
                        <td>{{ cantidad }}</td>
                        <td>{{ "%.2f"|format(total) }}</td>
                    </tr>
                    {% endfor %}
                    <tr class="total">
                        <td colspan="2">Total del Día</td>
                        <td>{{ "%.2f"|format(dia.total_dia) }}</td>
                    </tr>
                </tbody>
            </table>
        {% endfor %}
    </body>
    </html>
    """

    html_rendered = render_template_string(html_template, datos=datos_por_fecha)
    pdf = HTML(string=html_rendered).write_pdf()

    response = make_response(pdf)
    response.headers['Content-Type'] = 'application/pdf'
    response.headers['Content-Disposition'] = 'attachment; filename=reporte_ventas_detalladas_diarias.pdf'
    return response


############################################### Modulo Reportes #####################################################
@app.route('/reporte/excel/ventas-por-usuario', methods=['GET'])
def reporte_excel_ventas_por_usuario():
    fecha_limite = datetime.now() - timedelta(days=30)

    resultados = (
        db.session.query(
            Usuario.nombre,
            func.count(Venta.id_venta).label('cantidad_ventas'),
            func.sum(Venta.total).label('total_vendido')
        )
        .join(Venta, Usuario.id_usuario == Venta.id_usuario)
        .filter(Venta.fecha >= fecha_limite)
        .group_by(Usuario.nombre)
        .order_by(func.sum(Venta.total).desc())
        .all()
    )

    wb = Workbook()
    ws = wb.active
    ws.title = "Ventas por Usuario"

    ws.append(["Usuario", "Cantidad de Ventas", "Total Vendido (Q)"])

    for nombre, cantidad, total in resultados:
        ws.append([nombre, cantidad, float(total)])

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(
        output,
        as_attachment=True,
        download_name="reporte_ventas_usuarios.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

############################################### Modulo Reportes #####################################################
@app.route('/reporte/excel/insumos-detallado-por-mes', methods=['GET'])
def reporte_excel_insumos_detallado_por_mes():
    # Obtenemos todos los insumos ordenados por mes
    resultados = (
        db.session.query(
            Inventario.nombre,
            Inventario.cantidad,
            Inventario.unidad,
            Inventario.fecha_actualizacion
        )
        .order_by(Inventario.fecha_actualizacion)
        .all()
    )

    # Creamos libro Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Insumos por Mes"

    # Encabezado
    ws.append(["Mes", "Nombre del Insumo", "Cantidad", "Unidad"])

    # Datos
    for nombre, cantidad, unidad, fecha in resultados:
        mes = fecha.strftime('%Y-%m')
        ws.append([mes, nombre, cantidad, unidad])

    # Guardar y enviar
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(
        output,
        as_attachment=True,
        download_name="reporte_insumos_por_mes_detallado.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )


if __name__ == "__main__":
    with app.app_context():
        db.create_all()  
    app.run(debug=True)